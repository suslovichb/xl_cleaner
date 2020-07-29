from tkinter import Tk, Text, StringVar, BooleanVar, _setit, messagebox, filedialog
from tkinter.ttk import Style, Label, Button, OptionMenu, Checkbutton, Progressbar
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Side
from openpyxl.styles.borders import Border
import re
from os import path, system

# ws = load_workbook(filename, data_only=True)
wb = load_workbook('test.xlsx')
ws = wb['Лист1']

print(ws[1])
