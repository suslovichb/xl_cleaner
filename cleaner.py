from tkinter import Tk, Text, StringVar, BooleanVar, _setit, messagebox, filedialog
from tkinter.ttk import Style, Label, Button, OptionMenu, Checkbutton, Progressbar
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Side
from openpyxl.styles.borders import Border
import re
from os import path, system

REPLACEMENTS_PATH = "replacements.txt"
SORT_SETTINGS_PATH = "sorting.txt"
CATEGORY_ID_COLUMN_NAME = 'ID'

codes_dict = {
'&#09;': ' ',
'&#10;': ' ',
'&#13;': ' ',
'&#160;': ' ',
'&nbsp;': ' ',
'&#32;': ' ',
'&ndash;': '-',
'&#33;': '!',
'&#34;': '"',
'&quot;': '"',
'&#35;': '#',
'&#36;': '$',
'&#37;': '%',
'&amp;amp;': '&',
'&#38;': '&',
'&amp;': '&',
'&#39;': "'",
'&apos;': "'",
'&#40;': '(',
'&#41;': ')',
'&#42;': '*',
'&#43;': '+',
'&#44;': ',',
'&#45;': '-',
'&#46;': '.',
'&#47;': '/',
'&#48;': '0',
'&#49;': '1',
'&#50;': '2',
'&#51;': '3',
'&#52;': '4',
'&#53;': '5',
'&#54;': '6',
'&#55;': '7',
'&#56;': '8',
'&#57;': '9',
'&#58;': ':',
'&#59;': ';',
'&#60;': '<',
'&lt;': '<',
'&#61;': '=',
'&#62;': '>',
'&gt;': '>',
'&#63;': '?',
'&#64;': '@',
'&#65;': 'A',
'&#66;': 'B',
'&#67;': 'C',
'&#68;': 'D',
'&#69;': 'E',
'&#70;': 'F',
'&#71;': 'G',
'&#72;': 'H',
'&#73;': 'I',
'&#74;': 'J',
'&#75;': 'K',
'&#76;': 'L',
'&#77;': 'M',
'&#78;': 'N',
'&#79;': 'O',
'&#80;': 'P',
'&#81;': 'Q',
'&#82;': 'R',
'&#83;': 'S',
'&#84;': 'T',
'&#85;': 'U',
'&#86;': 'V',
'&#87;': 'W',
'&#88;': 'X',
'&#89;': 'Y',
'&#90;': 'Z',
'&#91;': '[',
'&#92;': '\\',
'&#93;': ']',
'&#94;': '^',
'&#95;': '_',
'&#96;': '`',
'&#97;': 'a',
'&#98;': 'b',
'&#99;': 'c',
'&#100;': 'd',
'&#101;': 'e',
'&#102;': 'f',
'&#103;': 'g',
'&#104;': 'h',
'&#105;': 'i',
'&#106;': 'j',
'&#107;': 'k',
'&#108;': 'l',
'&#109;': 'm',
'&#110;': 'n',
'&#111;': 'o',
'&#112;': 'p',
'&#113;': 'q',
'&#114;': 'r',
'&#115;': 's',
'&#116;': 't',
'&#117;': 'u',
'&#118;': 'v',
'&#119;': 'w',
'&#120;': 'x',
'&#121;': 'y',
'&#122;': 'z',
'&#123;': '{',
'&#124;': '|',
'&#125;': '}',
'&#126;': '~',
'&#161;': '¡',
'&iexcl;': '¡',
'&#162;': '¢',
'&cent;': '¢',
'&#163;': '£',
'&pound;': '£',
'&#164;': '¤',
'&curren;': '¤',
'&#165;': '¥',
'&yen;': '¥',
'&#166;': '¦',
'&brvbar;': '¦',
'&#167;': '§',
'&sect;': '§',
'&#168;': '¨',
'&uml;': '¨',
'&#169;': '©',
'&copy;': '©',
'&#170;': 'ª',
'&ordf;': 'ª',
'&#171;': '«',
'&laquo;': '«',
'&#172;': '¬',
'&not;': '¬',
'&#173;': '­',
'&shy;': '­',
'&#175;': '¯',
'&macr;': '¯',
'&#176;': '°',
'&deg;': '°',
'&#177;': '±',
'&plusmn;': '±',
'&#178;': '²',
'&sup2;': '²',
'&#179;': '³',
'&sup3;': '³',
'&#180;': '´',
'&acute;': '´',
'&#181;': 'µ',
'&micro;': 'µ',
'&#182;': '¶',
'&para;': '¶',
'&#183;': '·',
'&middot;': '·',
'&#184;': '¸',
'&cedil;': '¸',
'&#185;': '¹',
'&sup1;': '¹',
'&#186;': 'º',
'&ordm;': 'º',
'&#187;': '»',
'&raquo;': '»',
'&#188;': '¼',
'&frac14;': '¼',
'&#189;': '½',
'&frac12;': '½',
'&#190;': '¾',
'&frac34;': '¾',
'&#191;': '¿',
'&iquest;': '¿',
'&#192;': 'À',
'&Agrave;': 'À',
'&#193;': 'Á',
'&Aacute;': 'Á',
'&#194;': 'Â',
'&Acirc;': 'Â',
'&#195;': 'Ã',
'&Atilde;': 'Ã',
'&#196;': 'Ä',
'&Auml;': 'Ä',
'&#197;': 'Å',
'&Aring;': 'Å',
'&#198;': 'Æ',
'&AElig;': 'Æ',
'&#199;': 'Ç',
'&Ccedil;': 'Ç',
'&#200;': 'È',
'&Egrave;': 'È',
'&#201;': 'É',
'&Eacute;': 'É',
'&#202;': 'Ê',
'&Ecirc;': 'Ê',
'&#203;': 'Ë',
'&Euml;': 'Ë',
'&#204;': 'Ì',
'&Igrave;': 'Ì',
'&#205;': 'Í',
'&Iacute;': 'Í',
'&#206;': 'Î',
'&Icirc;': 'Î',
'&#207;': 'Ï',
'&Iuml;': 'Ï',
'&#208;': 'Ð',
'&ETH;': 'Ð',
'&#209;': 'Ñ',
'&Ntilde;': 'Ñ',
'&#210;': 'Ò',
'&Ograve;': 'Ò',
'&#211;': 'Ó',
'&Oacute;': 'Ó',
'&#212;': 'Ô',
'&Ocirc;': 'Ô',
'&#213;': 'Õ',
'&Otilde;': 'Õ',
'&#214;': 'Ö',
'&Ouml;': 'Ö',
'&#215;': '×',
'&times;': '×',
'&#216;': 'Ø',
'&Oslash;': 'Ø',
'&#217;': 'Ù',
'&Ugrave;': 'Ù',
'&#218;': 'Ú',
'&Uacute;': 'Ú',
'&#219;': 'Û',
'&Ucirc;': 'Û',
'&#220;': 'Ü',
'&Uuml;': 'Ü',
'&#221;': 'Ý',
'&Yacute;': 'Ý',
'&#222;': 'Þ',
'&THORN;': 'Þ',
'&#223;': 'ß',
'&szlig;': 'ß',
'&#224;': 'à',
'&agrave;': 'à',
'&#225;': 'á',
'&aacute;': 'á',
'&#226;': 'â',
'&;': 'â',
'&#227;': 'ã',
'&atilde;': 'ã',
'&#228;': 'ä',
'&auml;': 'ä',
'&#229;': 'å',
'&aring;': 'å',
'&#230;': 'æ',
'&aelig;': 'æ',
'&#231;': 'ç',
'&ccedil;': 'ç',
'&#232;': 'è',
'&egrave;': 'è',
'&#233;': 'é',
'&eacute;': 'é',
'&#234;': 'ê',
'&ecirc;': 'ê',
'&#235;': 'ë',
'&euml;': 'ë',
'&#236;': 'ì',
'&igrave;': 'ì',
'&#237;': 'í',
'&iacute;': 'í',
'&#238;': 'î',
'&icirc;': 'î',
'&#239;': 'ï',
'&iuml;': 'ï',
'&#240;': 'ð',
'&eth;': 'ð',
'&#241;': 'ñ',
'&ntilde;': 'ñ',
'&#242;': 'ò',
'&ograve;': 'ò',
'&#243;': 'ó',
'&oacute;': 'ó',
'&#244;': 'ô',
'&ocirc;': 'ô',
'&#245;': 'õ',
'&otilde;': 'õ',
'&#246;': 'ö',
'&ouml;': 'ö',
'&#247;': '÷',
'&divide;': '÷',
'&#248;': 'ø',
'&oslash;': 'ø',
'&#249;': 'ù',
'&ugrave;': 'ù',
'&#250;': 'ú',
'&uacute;': 'ú',
'&#251;': 'û',
'&ucirc;': 'û',
'&#252;': 'ü',
'&uuml;': 'ü',
'&#253;': 'ý',
'&yacute;': 'ý',
'&#254;': 'þ',
'&thorn;': 'þ',
'&#255;': 'ÿ',
'&yuml;': 'ÿ',
'&#38;': '&',
'&amp;': '&',
'&#8226;': '•',
'&bull;': '•',
'&#9702;': '◦',
'&#8729;': '∙',
'&#8227;': '‣',
'&#8259;': '⁃',
'&#176;': '°',
'&deg;': '°',
'&#8734;': '∞',
'&infin;': '∞',
'&#8240;': '‰',
'&permil;': '‰',
'&#8901;': '⋅',
'&sdot;': '⋅',
'&#177;': '±',
'&plusmn;': '±',
'&#8224;': '†',
'&dagger;': '†',
'&#8212;': '—',
'&mdash;': '—',
'&#172;': '¬',
'&not;': '¬',
'&#181;': 'µ',
'&micro;': 'µ',
'&#8869;': '⊥',
'&perp;': '⊥',
'&#8741;': '∥',
'&par;': '∥',
'&#36;': '$',
'&#8364;': '€',
'&euro;': '€',
'&#163;': '£',
'&pound;': '£',
'&#165;': '¥',
'&yen;': '¥',
'&#162;': '¢',
'&cent;': '¢',
'&#8377;': '₹',
'&#8360;': '₨',
'&#8369;': '₱',
'&#8361;': '₩',
'&#3647;': '฿',
'&#8363;': '₫',
'&#8362;': '₪',
'&#169;': '©',
'&copy;': '©',
'&#174;': '®',
'&reg;':'®',
'&#8471;': '℗',
'&#8482;': '™',
'&trade;': '™',
'&#8480;': '℠',
'&#945;': 'α',
'&alpha;': 'α',
'&#946;': 'β',
'&beta;': 'β',
'&#947;': 'γ',
'&gamma;': 'γ',
'&#948;': 'δ',
'&delta;': 'δ',
'&#949;': 'ε',
'&epsilon;': 'ε',
'&#950;': 'ζ',
'&zeta;': 'ζ',
'&#951;': 'η',
'&eta;': 'η',
'&#952;': 'θ',
'&theta;': 'θ',
'&#953;': 'ι',
'&iota;': 'ι',
'&#954;': 'κ',
'&kappa;': 'κ',
'&#955;': 'λ',
'&lambda;': 'λ',
'&#956;': 'μ',
'&mu;': 'μ',
'&#957;': 'ν',
'&nu;': 'ν',
'&#958;': 'ξ',
'&xi;': 'ξ',
'&#959;': 'ο',
'&omicron;': 'ο',
'&#960;': 'π',
'&pi;': 'π',
'&#961;': 'ρ',
'&rho;': 'ρ',
'&#963;': 'σ',
'&sigma;': 'σ',
'&#964;': 'τ',
'&tau;': 'τ',
'&#965;': 'υ',
'&upsilon;': 'υ',
'&#966;': 'φ',
'&phi;': 'φ',
'&#967;': 'χ',
'&chi;': 'χ',
'&#968;': 'ψ',
'&psi;': 'ψ',
'&#969;': 'ω',
'&omega;': 'ω',
'&#913;': 'Α',
'&Alpha;': 'Α',
'&#914;': 'Β',
'&Beta;': 'Β',
'&#915;': 'Γ',
'&Gamma;': 'Γ',
'&#916;': 'Δ',
'&Delta;': 'Δ',
'&#917;': 'Ε',
'&Epsilon;': 'Ε',
'&#918;': 'Ζ',
'&Zeta;': 'Ζ',
'&#919;': 'Η',
'&Eta;': 'Η',
'&#920;': 'Θ',
'&Theta;': 'Θ',
'&#921;': 'Ι',
'&Iota;': 'Ι',
'&#922;': 'Κ',
'&Kappa;': 'Κ',
'&#923;': 'Λ',
'&Lambda;': 'Λ',
'&#924;': 'Μ',
'&Mu;': 'Μ',
'&#925;': 'Ν',
'&Nu;': 'Ν',
'&#926;': 'Ξ',
'&Xi;': 'Ξ',
'&#927;': 'Ο',
'&Omicron;': 'Ο',
'&#928;': 'Π',
'&Pi;': 'Π',
'&#929;': 'Ρ',
'&Rho;': 'Ρ',
'&#931;': 'Σ',
'&Sigma;': 'Σ',
'&#932;': 'Τ',
'&Tau;': 'Τ',
'&#933;': 'Υ',
'&Upsilon;': 'Υ',
'&#934;': 'Φ',
'&Phi;': 'Φ',
'&#935;': 'Χ',
'&Chi;': 'Χ',
'&#936;': 'Ψ',
'&Psi;': 'Ψ',
'&#937;': 'Ω',
'&Omega;': 'Ω',
'&#09': ' ',
# '&#10': ' ',
'&#13': ' ',
'&#160': ' ',
'&nbsp': ' ',
'&#32': ' ',
'&#33': '!',
'&#34': '"',
'&quot': '"',
'&#35': '#',
# '&#36': '$',
'&#37': '%',
'&#38': '&',
'&amp': '&',
'&#39': "'",
'&apos': "'",
'&#40': '(',
'&#41': ')',
'&#42': '*',
'&#43': '+',
'&#44': ',',
'&#45': '-',
'&#46': '.',
'&#47': '/',
'&#48': '0',
'&#49': '1',
'&#50': '2',
'&#51': '3',
'&#52': '4',
'&#53': '5',
'&#54': '6',
'&#55': '7',
'&#56': '8',
'&#57': '9',
'&#58': ':',
'&#59': ';',
'&#60': '<',
'&lt': '<',
'&#61': '=',
'&#62': '>',
'&gt': '>',
'&#63': '?',
'&#64': '@',
'&#65': 'A',
'&#66': 'B',
'&#67': 'C',
'&#68': 'D',
'&#69': 'E',
'&#70': 'F',
'&#71': 'G',
'&#72': 'H',
'&#73': 'I',
'&#74': 'J',
'&#75': 'K',
'&#76': 'L',
'&#77': 'M',
'&#78': 'N',
'&#79': 'O',
'&#80': 'P',
'&#81': 'Q',
# '&#82': 'R',
# '&#83': 'S',
# '&#84': 'T',
'&#85': 'U',
'&#86': 'V',
# '&#87': 'W',
# '&#88': 'X',
# '&#89': 'Y',
'&#90': 'Z',
# '&#91': '[',
# '&#92': '\\',
# '&#93': ']',
# '&#94': '^',
# '&#95': '_',
# '&#96': '`',
# '&#97': 'a',
'&#98': 'b',
'&#99': 'c',
'&#100': 'd',
'&#101': 'e',
'&#102': 'f',
'&#103': 'g',
'&#104': 'h',
'&#105': 'i',
'&#106': 'j',
'&#107': 'k',
'&#108': 'l',
'&#109': 'm',
'&#110': 'n',
'&#111': 'o',
'&#112': 'p',
'&#113': 'q',
'&#114': 'r',
'&#115': 's',
'&#116': 't',
'&#117': 'u',
'&#118': 'v',
'&#119': 'w',
'&#120': 'x',
'&#121': 'y',
'&#122': 'z',
'&#123': '{',
'&#124': '|',
'&#125': '}',
'&#126': '~',
'&#161': '¡',
'&iexcl': '¡',
'&#162': '¢',
'&cent': '¢',
'&#163': '£',
'&pound': '£',
'&#164': '¤',
'&curren': '¤',
'&#165': '¥',
'&yen': '¥',
'&#166': '¦',
'&brvbar': '¦',
'&#167': '§',
'&sect': '§',
'&#168': '¨',
'&uml': '¨',
'&#169': '©',
'&copy': '©',
'&#170': 'ª',
'&ordf': 'ª',
'&#171': '«',
'&laquo': '«',
'&#172': '¬',
'&not': '¬',
'&#173': '­',
'&shy': '­',
'&#174': '®',
'&reg': '®',
'&#175': '¯',
'&macr': '¯',
'&#176': '°',
'&deg': '°',
'&#177': '±',
'&plusmn': '±',
'&#178': '²',
'&sup2': '²',
'&#179': '³',
'&sup3': '³',
'&#180': '´',
'&acute': '´',
'&#181': 'µ',
'&micro': 'µ',
'&#182': '¶',
'&para': '¶',
'&#183': '·',
'&middot': '·',
'&#184': '¸',
'&cedil': '¸',
'&#185': '¹',
'&sup1': '¹',
'&#186': 'º',
'&ordm': 'º',
'&#187': '»',
'&raquo': '»',
'&#188': '¼',
'&frac14': '¼',
'&#189': '½',
'&frac12': '½',
'&#190': '¾',
'&frac34': '¾',
'&#191': '¿',
'&iquest': '¿',
'&#192': 'À',
'&Agrave': 'À',
'&#193': 'Á',
'&Aacute': 'Á',
'&#194': 'Â',
'&Acirc': 'Â',
'&#195': 'Ã',
'&Atilde': 'Ã',
'&#196': 'Ä',
'&Auml': 'Ä',
'&#197': 'Å',
'&Aring': 'Å',
'&#198': 'Æ',
'&AElig': 'Æ',
'&#199': 'Ç',
'&Ccedil': 'Ç',
'&#200': 'È',
'&Egrave': 'È',
'&#201': 'É',
'&Eacute': 'É',
'&#202': 'Ê',
'&Ecirc': 'Ê',
'&#203': 'Ë',
'&Euml': 'Ë',
'&#204': 'Ì',
'&Igrave': 'Ì',
'&#205': 'Í',
'&Iacute': 'Í',
'&#206': 'Î',
'&Icirc': 'Î',
'&#207': 'Ï',
'&Iuml': 'Ï',
'&#208': 'Ð',
'&ETH': 'Ð',
'&#209': 'Ñ',
'&Ntilde': 'Ñ',
'&#210': 'Ò',
'&Ograve': 'Ò',
'&#211': 'Ó',
'&Oacute': 'Ó',
'&#212': 'Ô',
'&Ocirc': 'Ô',
'&#213': 'Õ',
'&Otilde': 'Õ',
'&#214': 'Ö',
'&Ouml': 'Ö',
'&#215': '×',
'&times': '×',
'&#216': 'Ø',
'&Oslash': 'Ø',
'&#217': 'Ù',
'&Ugrave': 'Ù',
'&#218': 'Ú',
'&Uacute': 'Ú',
'&#219': 'Û',
'&Ucirc': 'Û',
'&#220': 'Ü',
'&Uuml': 'Ü',
'&#221': 'Ý',
'&Yacute': 'Ý',
'&#222': 'Þ',
'&THORN': 'Þ',
'&#223': 'ß',
'&szlig': 'ß',
'&#224': 'à',
'&agrave': 'à',
'&#225': 'á',
'&aacute': 'á',
'&#226': 'â',
# '&': 'â',
'&#227': 'ã',
'&atilde': 'ã',
'&#228': 'ä',
'&auml': 'ä',
'&#229': 'å',
'&aring': 'å',
'&#230': 'æ',
'&aelig': 'æ',
'&#231': 'ç',
'&ccedil': 'ç',
'&#232': 'è',
'&egrave': 'è',
'&#233': 'é',
'&eacute': 'é',
'&#234': 'ê',
'&ecirc': 'ê',
'&#235': 'ë',
'&euml': 'ë',
'&#236': 'ì',
'&igrave': 'ì',
'&#237': 'í',
'&iacute': 'í',
'&#238': 'î',
'&icirc': 'î',
'&#239': 'ï',
'&iuml': 'ï',
'&#240': 'ð',
'&eth': 'ð',
'&#241': 'ñ',
'&ntilde': 'ñ',
'&#242': 'ò',
'&ograve': 'ò',
'&#243': 'ó',
'&oacute': 'ó',
'&#244': 'ô',
'&ocirc': 'ô',
'&#245': 'õ',
'&otilde': 'õ',
'&#246': 'ö',
'&ouml': 'ö',
'&#247': '÷',
'&divide': '÷',
'&#248': 'ø',
'&oslash': 'ø',
'&#249': 'ù',
'&ugrave': 'ù',
'&#250': 'ú',
'&uacute': 'ú',
'&#251': 'û',
'&ucirc': 'û',
'&#252': 'ü',
'&uuml': 'ü',
'&#253': 'ý',
'&yacute': 'ý',
'&#254': 'þ',
'&thorn': 'þ',
'&#255': 'ÿ',
'&yuml': 'ÿ',
'&#38': '&',
'&amp': '&',
'&#8226': '•',
'&bull': '•',
'&#9702': '◦',
'&#8729': '∙',
'&#8227': '‣',
'&#8259': '⁃',
'&#176': '°',
'&deg': '°',
'&#8734': '∞',
'&infin': '∞',
'&#8240': '‰',
'&permil': '‰',
'&#8901': '⋅',
'&sdot': '⋅',
'&#177': '±',
'&plusmn': '±',
'&#8224': '†',
'&dagger': '†',
'&#8212': '—',
'&mdash': '—',
'&#172': '¬',
'&not': '¬',
'&#181': 'µ',
'&micro': 'µ',
'&#8869': '⊥',
'&perp': '⊥',
'&#8741': '∥',
'&par': '∥',
'&#36': '$',
'&#8364': '€',
'&euro': '€',
'&#163': '£',
'&pound': '£',
'&#165': '¥',
'&yen': '¥',
'&#162': '¢',
'&cent': '¢',
'&#8377': '₹',
'&#8360': '₨',
'&#8369': '₱',
'&#8361': '₩',
'&#3647': '฿',
'&#8363': '₫',
'&#8362': '₪',
'&#169': '©',
'&copy': '©',
'&#174': '®',
'&#8471': '℗',
'&#8482': '™',
'&trade': '™',
'&#8480': '℠',
'&#945': 'α',
'&alpha': 'α',
'&#946': 'β',
'&beta': 'β',
'&#947': 'γ',
'&gamma': 'γ',
'&#948': 'δ',
'&delta': 'δ',
'&#949': 'ε',
'&epsilon': 'ε',
'&#950': 'ζ',
'&zeta': 'ζ',
'&#951': 'η',
'&eta': 'η',
'&#952': 'θ',
'&theta': 'θ',
'&#953': 'ι',
'&iota': 'ι',
'&#954': 'κ',
'&kappa': 'κ',
'&#955': 'λ',
'&lambda': 'λ',
'&#956': 'μ',
'&mu': 'μ',
'&#957': 'ν',
'&nu': 'ν',
'&#958': 'ξ',
'&xi': 'ξ',
'&#959': 'ο',
'&omicron': 'ο',
'&#960': 'π',
'&pi': 'π',
'&#961': 'ρ',
'&rho': 'ρ',
'&#963': 'σ',
'&sigma': 'σ',
'&#964': 'τ',
'&tau': 'τ',
'&#965': 'υ',
'&upsilon': 'υ',
'&#966': 'φ',
'&phi': 'φ',
'&#967': 'χ',
'&chi': 'χ',
'&#968': 'ψ',
'&psi': 'ψ',
'&#969': 'ω',
'&omega': 'ω',
'&#913': 'Α',
'&Alpha': 'Α',
'&#914': 'Β',
'&Beta': 'Β',
'&#915': 'Γ',
'&Gamma': 'Γ',
'&#916': 'Δ',
'&Delta': 'Δ',
'&#917': 'Ε',
'&Epsilon': 'Ε',
'&#918': 'Ζ',
'&Zeta': 'Ζ',
'&#919': 'Η',
'&Eta': 'Η',
'&#920': 'Θ',
'&Theta': 'Θ',
'&#921': 'Ι',
'&Iota': 'Ι',
'&#922': 'Κ',
'&Kappa': 'Κ',
'&#923': 'Λ',
'&Lambda': 'Λ',
'&#924': 'Μ',
'&Mu': 'Μ',
'&#925': 'Ν',
'&Nu': 'Ν',
'&#926': 'Ξ',
'&Xi': 'Ξ',
'&#927': 'Ο',
'&Omicron': 'Ο',
'&#928': 'Π',
'&Pi': 'Π',
'&#929': 'Ρ',
'&Rho': 'Ρ',
'&#931': 'Σ',
'&Sigma': 'Σ',
'&#932': 'Τ',
'&Tau': 'Τ',
'&#933': 'Υ',
'&Upsilon': 'Υ',
'&#934': 'Φ',
'&Phi': 'Φ',
'&#935': 'Χ',
'&Chi': 'Χ',
'&#936': 'Ψ',
'&Psi': 'Ψ',
'&#937': 'Ω',
'&Omega': 'Ω',
# '&#117;': 'u',
'&ldquo;': '"',
'&ldquo': '"',
'&rdquo;': '"',
'&rdquo': '"',
'&minus;': '-',
'&minus': '-',
'&helip;': '…',
'&helip': '…',
'&rsquo;': "'",
'&rsquo': "'",
'&lsquo;': "'",
'&lsquo': "'",
'&ndash;': '–',
'&ndash': '–'
}

sku_letters_dict = {
"," : "//0044",
'A' : '//0065',
'B' : '//0066',
'C' : '//0067',
'D' : '//0068',
'E' : '//0069',
'F' : '//0070',
'G' : '//0071',
'H' : '//0072',
'I' : '//0073',
'J' : '//0074',
'K' : '//0075',
'L' : '//0076',
'M' : '//0077',
'N' : '//0078',
'O' : '//0079',
'P' : '//0080',
'Q' : '//0081',
'R' : '//0082',
'S' : '//0083',
'T' : '//0084',
'U' : '//0085',
'V' : '//0086',
'W' : '//0087',
'X' : '//0088',
'Y' : '//0089',
'Z' : '//0090',
'a' : '//0097',
'b' : '//0098',
'c' : '//0099',
'd' : '//0100',
'e' : '//0101',
'f' : '//0102',
'g' : '//0103',
'h' : '//0104',
'i' : '//0105',
'j' : '//0106',
'k' : '//0107',
'l' : '//0108',
'm' : '//0109',
'n' : '//0110',
'o' : '//0111',
'p' : '//0112',
'q' : '//0113',
'r' : '//0114',
's' : '//0115',
't' : '//0116',
'u' : '//0117',
'v' : '//0118',
'w' : '//0119',
'x' : '//0120',
'y' : '//0121',
'z' : '//0122'
}


text_formatting_tags = [
'<b>',
'</b>',
'<strong>',
'</strong>',
'<i>',
'</i>',
'<em>',
'</em>',
'<u>',
'</u>',
'<tt>',
'</tt>',
'<s>',
'</s>',
'<big>',
'</big>',
'<small>',
'</small>',
'<font>',
'</font>',
'<center>',
'</center>',
'<sup>',
'</sup>',
'<sub>',
'</sub>'
]

table_tags = [
'<table>',
'</table>',
'<tr>',
'</tr>',
'<th>',
'</th>',
'<td>',
'</td>'
]

necessary_tags = [
'<br>',
'<h3>',
*text_formatting_tags,
*table_tags
]

def get_workbook():
    try:
        filename = workbook_field.get(1.0, 'end').replace('\n','')
        if format_sheet_check.get():
            return load_workbook(filename, data_only=True)
        return load_workbook(filename)
    except Exception:
        messagebox.showerror(title="Ошибка", message="Ошибка загрузки. Проверьте данные")
        enable_buttons()

def choose_file():
    choosed_file = filedialog.askopenfilename(filetypes=(("Excel files", "*.xlsx"),("All files","*.*")))
    if choosed_file:
        workbook_field.delete(1.0, 'end')
        workbook_field.insert(1.0, choosed_file)
        refresh_worksheets()

def refresh_worksheets():
    worksheet_choice.set("")
    worksheet_optionmenu['menu'].delete(0, 'end')
    wb = get_workbook()
    if not wb: return

    new_choices = wb.sheetnames
    for choice in new_choices:
        worksheet_optionmenu['menu'].add_command(label=choice, command=_setit(worksheet_choice, choice))

def worksheet_changed(*args):
    worksheet_field.delete(1.0, 'end')
    worksheet_field.insert(1.0, worksheet_choice.get())

def update_counter_text(current, total):
    counter_text.set('{}/{}'.format(current, total))

def disable_buttons():
    clean_button['state'] = 'disabled'
    open_button['state'] = 'disabled'
    edit_custom_button['state'] = 'disabled'
    edit_sort_button['state'] = 'disabled'

def enable_buttons():
    clean_button['state'] = 'normal'
    open_button['state'] = 'normal'
    edit_custom_button['state'] = 'normal'
    edit_sort_button['state'] = 'normal'

def clean():
    disable_buttons()
    progress['value'] = 0
    process_counter = 0
    processes_number = 0
    sheet_to_read = worksheet_field.get(1.0, 'end').replace('\n','')
    wb = get_workbook()
    if not wb: return

    try:
        ws = wb[sheet_to_read]
    except Exception:
        messagebox.showerror(title="Ошибка", message="Ошибка загрузки. Проверьте данные")
        enable_buttons()
        return

    if delete_columns_check.get(): processes_number += 1
    if replace_codes_check.get(): processes_number += 1
    if delete_tags_check.get(): processes_number += 1
    if replace_custom_check.get(): processes_number += 1
    if replace_letters_check.get(): processes_number += 1
    if count_amounts_in_categories_check.get(): processes_number += 1
    if sort_columns_check.get(): processes_number += 1
    if format_sheet_check.get(): processes_number += 1

    if processes_number == 0:
        messagebox.showerror(title="Ошибка", message="Выберите способы обратботки")
        enable_buttons()
        return

    update_counter_text(process_counter, processes_number)
    process_counter_label.grid(row=14, column=0, columnspan=5, pady=(20,5))

    if delete_columns_check.get():
        process_counter += 1
        update_counter_text(process_counter, processes_number)
        delete_columns(ws)

    if replace_codes_check.get():
        process_counter += 1
        update_counter_text(process_counter, processes_number)
        replace_codes(ws)

    if delete_tags_check.get():
        process_counter += 1
        update_counter_text(process_counter, processes_number)
        delete_tags(ws)

    if replace_custom_check.get():
        process_counter += 1
        update_counter_text(process_counter, processes_number)
        replace_custom(ws)

    if replace_letters_check.get():
        process_counter += 1
        update_counter_text(process_counter, processes_number)
        replace_sku_letters(ws)

    if count_amounts_in_categories_check.get():
        process_counter += 1
        update_counter_text(process_counter, processes_number)
        count_amounts_in_categories(ws)

    if sort_columns_check.get():
        process_counter += 1
        update_counter_text(process_counter, processes_number)
        sort_columns(ws)

    if format_sheet_check.get():
        process_counter += 1
        update_counter_text(process_counter, processes_number)
        format_sheet(ws)

    try:
        save_workbook(wb)
    except PermissionError:
        messagebox.showerror(title="Ошибка", message='Ошибка доступа')
    else:
        print("Done!")
        messagebox.showinfo(title="Очистка", message="Готово!")
    finally:
        progress['value'] = 0
        progress.update()
        process_counter_label.grid_forget()
        enable_buttons()
        return

def find_col_index(ws, col_name):
    for cell in ws[1]:
        if cell.value == col_name:
            return cell.col_idx

def replace_sku_letters(ws):
    progress['maximum'] = ws.max_row - 2
    progress['value'] = 0
    progress.update()
    col_num = find_col_index(ws, 'sku')
    if not col_num:
        messagebox.showerror(title="Ошибка", message="Столбец sku не найден")
        return
    for row_num in range(2, ws.max_row):
        cell = ws.cell(row_num, col_num)
        if cell.value:
            for letter, code in sku_letters_dict.items():
                cell.value = str(cell.value).replace(letter, code)
        progress['value'] += 1
        progress.update()

def get_sequences(list_of_ints):
    progress['value'] = 0
    progress.update()
    sequence_count = 1
    sequences = []
    for item in list_of_ints:
        next_item = None
        if list_of_ints.index(item) < (len(list_of_ints) - 1):
            next_item = list_of_ints[list_of_ints.index(item) + 1]

        if (item + 1) == next_item:
            sequence_count += 1
        else:
            first_in_sequence = list_of_ints[list_of_ints.index(item) - sequence_count + 1]
            sequences.append([first_in_sequence, sequence_count])
            sequence_count = 1

    return sequences

def delete_columns(ws):
    progress['value'] = 0
    progress.update()
    empty_cols_indices = []
    for index, column in list(enumerate(ws.iter_cols(values_only=True), start=1)):
        if not any(column[1:]) or column[0]==None:
            empty_cols_indices.append(index)

    empty_cols_sequences = get_sequences(empty_cols_indices)
    progress["maximum"] = len(empty_cols_sequences)

    for sequence in reversed(empty_cols_sequences):
        ws.delete_cols(sequence[0], sequence[1])
        progress['value'] += 1
        progress.update()

def replace_codes(ws):
    progress['maximum'] = ws.max_row
    progress['value'] = 0
    progress.update()
    for row in ws.iter_rows():
        row = tuple([replace_in_cell(cell, codes_dict) for cell in row])
        progress['value'] += 1
        progress.update()

def replace_in_cell(cell, dict):
    for code, character in dict.items():
        if cell.value:
            cell.value = str(cell.value).replace(code, character)

def replace_custom(ws):
    progress['maximum'] = ws.max_row
    progress['value'] = 0
    progress.update()
    custom_dict = {}
    try:
        f = open('replacements.txt', 'r', encoding='utf8')
        text = f.read()
        arr = text.split('\n')
        f.close()
        for item in arr:
            if ' : ' in item:
                pair = item.split(' : ')
                custom_dict[pair[0]] = pair[1]
    except FileNotFoundError:
        messagebox.showerror(title="Замена", message="Файл замен не обнаружен")
        return
    except:
        messagebox.showerror(title="Ошибка", message="Ошибка настраиваемой замены")
        return

    if custom_dict:
        for row in ws.iter_rows():
            row = tuple([replace_in_cell(cell, custom_dict) for cell in row])
            progress['value'] += 1
            progress.update()
    else:
        messagebox.showerror(title="Замена", message="Пользовательские значения не обнаружены")

def delete_tags(ws):
    progress['value'] = 0
    progress['maximum'] = ws.max_row
    progress.update()
    for row in ws.iter_rows():
        row = tuple([delete_tags_in_cell(cell) for cell in row])
        progress['value'] += 1
        progress.update()

def delete_tags_in_cell(cell):
    for code, character in codes_dict.items():
        if cell.value:
            cell.value = re.sub('(?!{})(<.*?>)'.format('|'.join(necessary_tags)),'<br>',str(cell.value))
            cell.value = re.sub('<br><br>+','<br>',str(cell.value))

def save_workbook(wb):
    filename = workbook_field.get(1.0, 'end').replace('\n','')
    if rewrite_check.get():
        wb.save(filename)
    else:
        wb.save(filename.replace(".xl","_cleaned.xl"))

def toggle_edit_button():
    if replace_custom_check.get():
        edit_custom_button['state'] = 'normal'
    else:
        edit_custom_button['state'] = 'disabled'

def edit_custom():
    if not path.isfile(REPLACEMENTS_PATH):
        with open(REPLACEMENTS_PATH, 'w'): pass
    system('start ' + REPLACEMENTS_PATH)

def toggle_edit_sort_button():
    if sort_columns_check.get():
        edit_sort_button['state'] = 'normal'
    else:
        edit_sort_button['state'] = 'disabled'

def edit_sort():
    if not path.isfile(SORT_SETTINGS_PATH):
        with open(SORT_SETTINGS_PATH, 'w'): pass
    system('start ' + SORT_SETTINGS_PATH)

def append_column(ws, column):
    new_col_num = ws.max_column + 1
    for row_num, cell in enumerate(column, start=1):
        ws.cell(row=row_num, column=new_col_num, value=cell.value)

def sort_columns(ws):
    progress['maximum'] = 8
    progress['value'] = 0
    progress.update()

    settings = {}
    try:
        f = open(SORT_SETTINGS_PATH, 'r', encoding='utf8')
        text = f.read()
        text = text.replace('\n', '')
        ###############################################
        text = re.sub(r' *([,:;]) *', r'\g<1>', text)
        ###############################################
        arr = text.split(';')
        f.close()
        for item in arr:
            if ':' in item:
                pair = item.split(':',1)
                settings[pair[0]] = pair[1].split(',')
    except FileNotFoundError:
        messagebox.showerror(title="Ошибка", message="Файл сортировки не обнаружен")
        return
    except:
        messagebox.showerror(title="Ошибка", message="Ошибка сортировки")
        return

    try:
        ordered_attributes = settings['priority_attributes']
        trash_attributes = settings['trash_attributes']
    except:
        messagebox.showerror(title="Сортировка", message="Некорректные настройки сортировки")
        return

    progress['value'] = 1
    progress.update()

    initial_columns_quantity = ws.max_column
    processed_columns_indices = []

    for attribute in ordered_attributes:
        col_num = find_col_index(ws, attribute)
        if col_num:
            column = ws[get_column_letter(col_num)]
            append_column(ws, column)
            processed_columns_indices.append(col_num)
        else:
            ws.cell(row=1, column=ws.max_column+1, value=str(attribute))

    progress['value'] = 2
    progress.update()

    pic_headers = [[index, cell.value] for index, cell in enumerate(ws[1], start=1) if bool(re.search(r'^\d+_picture$', str(cell.value)))]
    pic_headers.sort(key=lambda x: x[1])
    for header in pic_headers:
        column = ws[get_column_letter(header[0])]
        append_column(ws, column)
        processed_columns_indices.append(header[0])

    progress['value'] = 3
    progress.update()

    trash_columns_indices = []
    for attribute in trash_attributes:
        col_num = find_col_index(ws, attribute)
        if col_num:
            trash_columns_indices.append(col_num)

    progress['value'] = 4
    progress.update()

    undefined_columns_indices = [index for index in range(1, initial_columns_quantity+1) if index not in processed_columns_indices+trash_columns_indices]
    for index in undefined_columns_indices:
        column = ws[get_column_letter(index)]
        append_column(ws, column)

    progress['value'] = 5
    progress.update()

    trash_header_cell = ws.cell(1, ws.max_column+1)
    trash_header_cell.value = 'МУСОР'
    trash_header_cell.fill = PatternFill(start_color='FF0000',
                                         end_color='FF0000',
                                         fill_type='solid')
    trash_header_cell.font = Font(color = 'FFFFFF')
    for cell in ws[get_column_letter(ws.max_column)][1:]:
        cell.value = '|'

    progress['value'] = 6
    progress.update()

    for index in trash_columns_indices:
        column = ws[get_column_letter(index)]
        append_column(ws, column)
        processed_columns_indices.append(index)

    progress['value'] = 7
    progress.update()

    ws.delete_cols(1, initial_columns_quantity)

    progress['value'] = 8
    progress.update()

def set_auto_filter(ws):
    ws.auto_filter.ref = 'A1:{}{}'.format(get_column_letter(ws.max_column),
                                          ws.max_row)

def format_sheet(ws):
    progress['value'] = 0
    progress['maximum'] = 6
    progress.update()

    ws.sheet_view.zoomScale = 90
    no_fill = PatternFill(fill_type=None)
    side = Side(border_style=None)
    no_border = Border(left=side, right=side, top=side, bottom=side)
    font = Font(name='Times New Roman',
                size=12,
                bold=False,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='000000')
    bold_font = Font(name='Times New Roman',
                size=12,
                bold=True,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='000000')
    alignment = Alignment(horizontal='general',
                          vertical='bottom',
                          text_rotation=0,
                          wrap_text=False,
                          shrink_to_fit=False,
                          indent=0)

    progress['value'] = 1
    progress.update()

    for col_num, column in enumerate(ws.iter_cols()):
        ws.column_dimensions[get_column_letter(col_num+1)].width = 12.9
        for cell in column:
            cell.font = font
            cell.alignment = alignment
            cell.hyperlink = None
            cell.fill = no_fill
            cell.border = no_border
            cell.number_format = 'General'

    progress['value'] = 2
    progress.update()

    for row_num in range(1, ws.max_row+1):
        ws.row_dimensions[row_num].height = 15.5

    progress['value'] = 3
    progress.update()

    set_auto_filter(ws)

    progress['value'] = 4
    progress.update()

    for cell in ws[1]:
        cell.font = bold_font

    trash_col_num = find_col_index(ws, 'МУСОР')
    if trash_col_num:
        trash_header_cell = ws.cell(1, trash_col_num)
        trash_header_cell.fill = PatternFill(start_color='FF0000',
                                             end_color='FF0000',
                                             fill_type='solid')
        trash_header_cell.font = Font(name='Times New Roman',
                                      size=12,
                                      bold=True,
                                      color='FFFFFF')
    progress['value'] = 5
    progress.update()

    ws.freeze_panes = 'A2'

    progress['value'] = 6
    progress.update()

def count_amounts_in_categories(ws):
    progress['maximum'] = 1
    progress['value'] = 0
    progress.update()
    id_col_num = find_col_index(ws, CATEGORY_ID_COLUMN_NAME)
    if id_col_num:
        amounts_of_categories = {}
        id_column = ws[get_column_letter(id_col_num)]
        progress['maximum'] = ws.max_row * 2 -2
        progress.update()
        for cell in id_column[1:]:
            progress['value'] += 1
            progress.update()
            key = str(cell.value)
            amounts_of_categories[key] = amounts_of_categories.get(key, 0) + 1
        ws.insert_cols(id_col_num+1, 1)
        amount_column = ws[get_column_letter(id_col_num+1)]
        amount_column[0].value = 'Количество'
        for row_num in range(1, ws.max_row):
            progress['value'] += 1
            progress.update()
            amount_column[row_num].value = amounts_of_categories[str(id_column[row_num].value)]
    else:
        progress['value'] = 1
        progress.update()
        messagebox.showerror(title="Ошибка", message="Стоблец с ID категорий не найден")
        return


window = Tk()
window.title('XL Cleaner')
window.geometry('500x500')
window.resizable(0, 0)

secondary_button_style = Style()
secondary_button_style.configure('secondary.TButton', font=('Calibri',8))

workbook_label = Label(window, text="Файл")
workbook_label.grid(row=0, column=0, columnspan=5, pady=(20,5))

workbook_field = Text(window, width=50, height = 1)
workbook_field.grid(row=1, column=0, columnspan=3, padx=(10,5))

open_button = Button(window, text="Открыть", command=choose_file, style='secondary.TButton')
open_button.grid(row=1, column=4)

worksheet_label = Label(window, text="Лист")
worksheet_label.grid(row=2, column=0, columnspan=5, pady=(10,5))

worksheet_field = Text(window, width=50, height = 1)
worksheet_field.grid(row=3, column=0, columnspan=3, padx=(10,5))

worksheet_choice = StringVar()
worksheet_choice.set("")
worksheet_choice.trace("w",worksheet_changed)
worksheet_optionmenu = OptionMenu(window, worksheet_choice)
worksheet_optionmenu.grid(row=3, column=4, sticky="wns")
s = Style()
s.configure("TMenubutton", foreground="#ffffff00", background="gray83")

delete_columns_check = BooleanVar()
delete_columns_check.set(False)
delete_columns_checkbutton = Checkbutton(window, variable=delete_columns_check, onvalue=True, offvalue=False, text='Удалить пустые столбцы')
delete_columns_checkbutton.grid(row=4, column=0, columnspan=5, pady=(20,5), padx=16, sticky="w")

replace_codes_check = BooleanVar()
replace_codes_check.set(False)
replace_codes_checkbutton = Checkbutton(window, variable=replace_codes_check, onvalue=True, offvalue=False, text='Заменить HTML коды символов')
replace_codes_checkbutton.grid(row=5, column=0, columnspan=5, pady=(5,5), padx=16, sticky="w")

delete_tags_check = BooleanVar()
delete_tags_check.set(False)
delete_tags_checkbutton = Checkbutton(window, variable=delete_tags_check, onvalue=True, offvalue=False, text='Удалить HTML теги')
delete_tags_checkbutton.grid(row=6, column=0, columnspan=5, pady=(5,5), padx=16, sticky="w")

replace_custom_check = BooleanVar()
replace_custom_check.set(False)
replace_custom_checkbutton = Checkbutton(window, variable=replace_custom_check, onvalue=True, offvalue=False, text='Настраиваемая замена', command=toggle_edit_button)
replace_custom_checkbutton.grid(row=7, column=0, columnspan=2, pady=(5,5), padx=16, sticky="w")

edit_custom_button = Button(window, text="Редактировать", command=edit_custom, state = 'disabled', style='secondary.TButton')
edit_custom_button.grid(row=7, column=1, columnspan=2, ipady=0, ipadx=0)

replace_letters_check = BooleanVar()
replace_letters_check.set(False)
replace_letters_checkbutton = Checkbutton(window, variable=replace_letters_check, onvalue=True, offvalue=False, text='Заменить буквы в sku')
replace_letters_checkbutton.grid(row=8, column=0, columnspan=5, pady=(5,5), padx=16, sticky="w")

count_amounts_in_categories_check = BooleanVar()
count_amounts_in_categories_check.set(False)
count_amounts_in_categories_checkbutton = Checkbutton(window, variable=count_amounts_in_categories_check, onvalue=True, offvalue=False, text='Подсчитать количества в категориях')
count_amounts_in_categories_checkbutton.grid(row=9, column=0, columnspan=5, pady=(5,5), padx=16, sticky="w")


sort_columns_check = BooleanVar()
sort_columns_check.set(False)
sort_columns_checkbutton = Checkbutton(window, variable=sort_columns_check, onvalue=True, offvalue=False, text='Сортировать столбцы', command=toggle_edit_sort_button)
sort_columns_checkbutton.grid(row=10, column=0, columnspan=2, pady=(5,5), padx=16, sticky="w")

edit_sort_button = Button(window, text="Настройки", command=edit_sort, state = 'disabled', style='secondary.TButton')
edit_sort_button.grid(row=10, column=1, columnspan=2, ipady=0, ipadx=0)

format_sheet_check = BooleanVar()
format_sheet_check.set(False)
format_sheet_checkbutton = Checkbutton(window, variable=format_sheet_check, onvalue=True, offvalue=False, text='Форматировать')
format_sheet_checkbutton.grid(row=11, column=0, columnspan=5, pady=(5,5), padx=16, sticky="w")

rewrite_check = BooleanVar()
rewrite_check.set(False)
rewrite_checkbutton = Checkbutton(window, variable=rewrite_check, onvalue=True, offvalue=False, text='Сохранить в исходный файл')
rewrite_checkbutton.grid(row=12, column=0, columnspan=5, pady=(5,20), padx=16, sticky="w")

clean_button = Button(window, text="ЗАПУСК", command=clean)
clean_button.grid(row=13, column=0, columnspan=5)

progress = Progressbar(window, orient='horizontal', length=490, mode = 'determinate')
progress['value'] = 0
progress.grid(row=14, column=0, columnspan=5, padx=5, pady=(20,5))

counter_text = StringVar()
counter_text.set('counter')
process_counter_label = Label(window, text="", font=("Arial", 10), textvariable=counter_text)

window.mainloop()
